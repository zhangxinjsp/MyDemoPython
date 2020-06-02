#!/usr/bin/env python
# -*- coding: utf-8 -*-

from xlsxwriter.workbook import Workbook
from enchant.checker import SpellChecker
import enchant
import openpyxl
import os
import shutil

# pip install pyenchant
# brew install enchant

os.chdir('/Users/zhangxin/Desktop/spell')

CHECK_LANGUAGE = 'en_US'

except_words = ['CTRS']
error_list = []


def read_from_excel(excel_name):
    shutil.copy(excel_name, 'result.xlsx')
    write_excel = Workbook('result.xlsx')
    red_format = write_excel.add_format({'color': 'red'})

    excel_book = openpyxl.load_workbook(excel_name)
    print("excel type ", type(excel_book))

    for sheet_name in excel_book.sheetnames:
        sheet = excel_book[sheet_name]
        write_sheet = write_excel.add_worksheet(name=sheet_name)

        total = sheet.max_column * sheet.max_row
        current_index = 0
        for column in range(1, sheet.max_column + 1):
            for row in range(1, sheet.max_row + 1):
                current_index += 1
                print('\r%s (Progress: %d%%)' % (sheet_name, (int(float(current_index) / float(total) * 100))), end=" ")
                text = sheet.cell(row=row, column=column).value
                check_sentence(text)

                write_to_new_excel(column - 1, row - 1, text, write_sheet, red_format)
        print('')
    excel_book.close()
    write_excel.close()


def check_sentence(text):
    error_list.clear()
    if text is None:
        return
    d = enchant.Dict(CHECK_LANGUAGE)

    checker = SpellChecker(CHECK_LANGUAGE)
    checker.set_text(text)
    if checker is None:
        return
    try:
        for err in checker:
            if err.word in except_words:
                continue
            if d.check(err.word.lower()) or d.check(err.word.upper()):
                continue
            error_list.append((err.wordpos, err.word))
    except Exception as e:
        # print('=============', end='    ')
        # print(e)
        error_list.clear()


def write_to_new_excel(column, row, text, sheet, red):
    """
    :param column:
    :param row:
    :param text:
    :param sheet:
    :param red:
    :return:
    """
    if len(error_list) == 0:
        # print('++++++++++++++write to excel: %d-%d' % (column, row))
        sheet.write(row, column, text)
    else:
        format_pairs = []
        start_index = 0
        for error in error_list:
            index = error[0]
            if text[start_index: index]:
                format_pairs.append(text[start_index: index])
            format_pairs.extend((red, error[1]))

            start_index = index + len(error[1])

        if text[start_index:]:
            format_pairs.append(text[start_index:])
        # print('**************write to excel: %d-%d _______%d' % (column, row, len(format_pairs)))
        # print(format_pairs)
        # print(error_list)
        if len(format_pairs) <= 2:
            format_pairs.append(' ')

        sheet.write_rich_string(row, column, *format_pairs)


if __name__ == '__main__':
    read_from_excel('test_0507.xlsx')
