#!/usr/bin/env python
# -*- coding: utf-8 -*-


# 没有需要下载  pip install openpyxl
import openpyxl

# 打开工作簿对象
excel = openpyxl.load_workbook('excel_test.xlsx')
print("excel type ", type(excel))

print(excel.sheetnames)
# 获取工作表对象，及工作表对象的标题
sheetName1 = excel.sheetnames[0]
sheet1 = excel[sheetName1]
print(sheet1.title)
print(sheet1)

# 获取活动工作表对象（上次操作的工作表）
print(excel.active)

# 从工作表中获取单元格对象
cell_a1 = sheet1["A1"]

print(cell_a1.value)  # 单元格的值
print(cell_a1.row)
print(cell_a1.column)
print(cell_a1.coordinate)

# 直接指定数字获取单元格
cell22 = sheet1.cell(row=2, column=2)
print(cell22.value)

# 获取最大行列信息
print(sheet1.max_row)
print(sheet1.max_column)

# 遍历
for i in range(1, sheet1.max_row + 1):
    for j in range(1, sheet1.max_column + 1):
        print(sheet1.cell(row=i, column=j).value, " ")
    print("")
